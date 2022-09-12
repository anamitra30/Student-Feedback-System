import tkinter as tk
from tkinter import ttk
from tkinter import Menu
from tkinter import messagebox as m_box
from tkinter import font
from tkinter import BOTH
from tkinter import filedialog
import mysql.connector as mysql
from openpyxl import load_workbook
from openpyxl import Workbook
from os import path, mkdir
from datetime import date
import admin_credentials


class Welcome:

    def __init__(self, master):
        self.master = master
        self.master.title("STUDENT FEEDBACK SYSTEM")
        self.master.geometry("480x420+500+150")
        self.master.resizable(False, False)

        # Fonts and Styles

        self.customized_font1 = font.Font(family='Helvetica', size=48, underline='True',
                                          weight="bold", slant='italic')
        self.customized_font2 = font.Font(family='Helvetica', size=34, slant='italic')
        self.style = ttk.Style()
        self.style.configure("buttons.TButton", foreground="#178c02", background='#332104',
                             font=('Helvetica', 15))

        # Creating Label frames

        self.labelframe1 = tk.LabelFrame(self.master, bd=0)
        self.labelframe1.pack(fill=BOTH, expand=1)

        self.labelframe2 = tk.LabelFrame(self.master, bd=0)
        self.labelframe2.pack(fill=BOTH, expand=1)

        # Creating a Menu Bar

        self.menu_bar = Menu(self.master)

        self.master.config(menu=self.menu_bar)

        self.file_menu = Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Exit", command=self._quit)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        self.help_menu = Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="About", command=self.about)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)

        # Creating Labels

        self.heading_label = tk.Label(self.labelframe1, text="RCCIIT", font=self.customized_font1)
        self.heading_label.grid(row=0, column=0, padx=40, pady=10, sticky="WE")

        self.info_label = tk.Label(self.labelframe1, text="Department of CSE\n(NBA Accredited)",
                                   font=self.customized_font2)
        self.info_label.grid(row=1, column=0, padx=40, pady=5, sticky="W")

        # Creating Buttons

        self.student_btn = ttk.Button(self.labelframe2, text="LOGIN AS STUDENT",
                                      width=25, style='buttons.TButton')
        self.student_btn.grid(row=0, column=0, sticky='w', padx=100, pady=15)

        self.faculty_btn = ttk.Button(self.labelframe2, text="LOGIN AS FACULTY",
                                      width=25, style='buttons.TButton')
        self.faculty_btn.grid(row=1, column=0, sticky='w', padx=100, pady=15)

        # Event Bind to the Buttons

        self.student_btn.bind('<Button-1>', self.student_login)
        self.student_btn.bind('<Return>', self.student_login)
        self.faculty_btn.bind('<Button-1>', self.faculty_login)
        self.faculty_btn.bind('<Return>', self.faculty_login)

    def _quit(self):
        self.master.quit()
        self.master.destroy()
        exit()

    def about(self):
        m_box.showinfo('ABOUT ', 'content.....')

    def student_login(self, event):
        root = tk.Tk()
        self.master.iconify()
        LoginStudent(root, self.master)

    def faculty_login(self, event):
        root = tk.Tk()
        self.master.iconify()
        LoginFaculty(root, self.master)


class LoginStudent:
    def __init__(self, this_window, previous_window):
        self.win = this_window
        self.prev_win = previous_window
        self.win.title("STUDENT LOGIN PAGE")
        self.win.geometry('+550+200')
        self.win.resizable(False, False)

        # -------- Fonts and Styles ------------------

        self.style = ttk.Style(self.win)

        self.style.configure("labels.TLabel", foreground="#964e06",
                             font=('Helvetica', 18, 'bold', 'italic'))

        self.style.configure("entry.TEntry", foreground='black')

        self.style.configure("buttons.TButton", foreground="#122e9b", background='black',
                             font=('Helvetica', 11, 'bold'))
        self.style.configure('buttons1.TButton', foreground='#5b022f', background='black',
                             font=('Helvetica', 10, 'bold', 'italic'))
        self.style.configure('buttons2.TButton', foreground='black', background='#4f2e68',
                             font=('Helvetica', 10, 'bold', 'italic'))

        # ------------------------- Creating a Menu Bar --------------------------

        self.menu_bar = Menu(self.win)

        self.win.config(menu=self.menu_bar)

        self.file_menu = Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Exit", command=self._quit)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        self.help_menu = Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="About", command=self.about)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)

        # ----------------------------- Creating Label frames -----------------------------------

        self.labelframe1 = ttk.LabelFrame(self.win)
        self.labelframe1.pack(fill=BOTH, expand=1)

        self.labelframe2 = ttk.LabelFrame(self.win)
        self.labelframe2.pack(fill=BOTH, expand=1)

        # ----------------------------------- Creating Labels -----------------------------------

        self.username = ttk.Label(self.labelframe1, text="USERNAME : ", style='labels.TLabel')
        self.username.grid(row=0, column=0, padx=30, pady=5, sticky="W")

        self.password = ttk.Label(self.labelframe1, text="PASSWORD : ", style='labels.TLabel')
        self.password.grid(row=1, column=0, padx=30, pady=5, sticky="W")

        self.year = ttk.Label(self.labelframe1, text="YEAR : ", style='labels.TLabel')
        self.year.grid(row=2, column=0, padx=30, pady=5, sticky="W")

        # ----------------------------------- Creating EntryFields ----------------------------

        self.name_entry = ttk.Entry(self.labelframe1, width=30, style='entry.TEntry')
        self.name_entry.grid(row=0, column=1, padx=20, pady=10)

        self.password_entry = ttk.Entry(self.labelframe1, width=30, show='*', style='entry.TEntry')
        self.password_entry.grid(row=1, column=1, padx=20, pady=10)

        # Creating ComboBox for the years

        self.combobox = ttk.Combobox(self.labelframe1, width=15,
                                     height=15, state='readonly')
        self.combobox['values'] = ['1ST', '2ND', '3RD', '4TH']
        self.combobox.grid(row=2, column=1, sticky='W', padx=20, pady=10)
        self.combobox.current(0)

        # Creating Buttons

        self.login_btn = ttk.Button(self.labelframe2, text="LOGIN ", width=15, style='buttons.TButton')
        self.login_btn.grid(row=0, column=0, sticky='w', padx=175, pady=8)

        self.hint_btn = ttk.Button(self.labelframe2, text="Hint", width=8, style='buttons1.TButton')
        self.hint_btn.grid(rows=1, column=0, sticky="w", padx=204, pady=2)

        self.go_to_homepage = ttk.Button(self.labelframe2, text="Home page \n    <<---",
                                         width=15, style='buttons2.TButton')
        self.go_to_homepage.grid(row=2, column=0, sticky='w', padx=20, pady=10)

        # Event Bind to the buttons

        self.login_btn.bind('<Button-1>', self.login_for_student)
        self.login_btn.bind('<Return>', self.login_for_student)
        self.hint_btn.bind('<Button-1>', self.hint)
        self.hint_btn.bind('<Return>', self.hint)
        self.go_to_homepage.bind('<Button-1>', self.prev_page)
        self.go_to_homepage.bind('<Return>', self.prev_page)

        self.name_entry.focus()

    def _quit(self):
        self.win.quit()
        self.win.destroy()
        exit()

    def about(self):
        m_box.showinfo('ABOUT ', 'content.....')

    def login_for_student(self, event):

        index = -1
        a = 1

        if len(self.name_entry.get()) == 0 or len(self.password_entry.get()) == 0:
            m_box.showerror('title....', 'Fill the form correctly')
        else:
            for i in range(len(rows)):
                if self.name_entry.get().upper().strip() == rows[i][0] and \
                        self.password_entry.get().upper().strip() == rows[i][1]:
                    if rows[i][3] != self.combobox.get().upper():
                        m_box.showwarning('', 'Fill the year correctly')
                        a = 0
                    else:
                        index = i
                        a = 1
                        break
            if index == -1 and a == 1:
                m_box.showerror('', 'Not Found!!')
            elif a == 1:
                if rows[index][4].upper() == 'ELIGIBLE':
                    if int(rows[index][5]) >= int(rows[index][6]):

                        self.prev_win.destroy()
                        self.win.destroy()

                        root = tk.Tk()
                        RatingPage(root, rows[index][0], rows[index][1],
                                   rows[index][2], rows[index][3])

                    else:
                        m_box.showerror('', 'Low Attendance')
                else:
                    m_box.showinfo('', 'Cannot submit multiple times')

    def prev_page(self, event):
        try:
            self.win.destroy()
            self.prev_win.deiconify()
        except tk.TclError:
            pass

    def hint(self, event):
        m_box.showinfo('HINT', 'NAME : YOUR FULL NAME ' + '\n\n' +
                       'PASSWORD : YOUR ROLL IN CSE20__/___ FORMAT')


class RatingPage:
    def __init__(self, this_window, student_name,
                 student_roll, student_section, student_year):

        self.window = this_window
        self.student_name = student_name
        self.student_roll = student_roll
        self.student_section = student_section
        self.student_year = student_year

        self.window.title("STUDENT FEEDBACK FORM")
        self.window.geometry("+100+150")
        self.window.resizable(False, False)

        # ----------------------- Fonts and Styles ----------------------

        self.design = ttk.Style(self.window)
        self.design.configure("Submit.TButton", foreground="red", background="black")

        # # ------------------ getting relevant details---------------

        # to store the data

        self.faculty = []
        self.paper_code = []
        self.paper_name = []

        if student_section == 'SEC-A':
            cursor.execute("select * from faculty_data_sec_a where year=%s", (self.student_year,))
        elif student_section == 'SEC-B':
            cursor.execute("select * from faculty_data_sec_b where year=%s", (self.student_year,))

        detail = cursor.fetchall()

        for i in range(len(detail)):
            self.faculty.append(detail[i][4])
            self.paper_code.append(detail[i][3])
            self.paper_name.append(detail[i][2])

        # Creating a Menu Bar

        self.menu_bar = Menu(self.window)
        self.window.config(menu=self.menu_bar)

        self.file_menu = Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Exit", command=self._quit)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        self.help_menu = Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="About", command=self.about)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)

        # --------------- ADD TABS -----------------

        self.tabs_arr = []

        self.tab_control = ttk.Notebook(self.window)  # create tab control

        for i in range(len(self.paper_code)):
            temp = ttk.Frame(self.tab_control)
            self.tab_control.add(temp, text=str(i + 1) + ". " + 'PAPER CODE :-\n    '
                                            + self.paper_code[i])
            self.tabs_arr.append(temp)

        for i in range(1):
            temp = ttk.Frame(self.tab_control)
            self.tab_control.add(temp, text=" FINAL\n PREVIEW   ")
            self.tabs_arr.append(temp)

        self.tab_control.pack(expand=1, fill="both")  # Pack to make visible

        # ------------------ To store the ratings ----------------------

        self.rating_knowledge = []
        self.rating_comm_skill = []
        self.rating_explain = []
        self.rating_commitment = []
        self.rating_interest_gen = []
        self.rating_qa = []
        self.rating_accessibility = []
        self.rating_tests = []
        self.rating_evaluation = []
        self.rating_overall = []

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_knowledge.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_comm_skill.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_explain.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_commitment.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_interest_gen.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_qa.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_accessibility.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_tests.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_evaluation.append(temp)

        for i in range(len(self.tabs_arr) - 1):
            temp = tk.IntVar(self.tabs_arr[i])
            temp.set(3)
            self.rating_overall.append(temp)

        # ----------------- Adding the Questions and other Labels --------------

        for i in range(len(self.tabs_arr) - 1):
            ttk.Label(self.tabs_arr[i], text='PAPER NAME :- ' + self.paper_name[i] + '\n' +
                                             'ASSIGNED FACULTY :- ' + self.faculty[i]). \
                grid(column=0, row=0, sticky='W', pady=12, padx=12)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'on the basis of Knowledge based teaching ? ') \
                .grid(row=1, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'on the basis of communication skill ? ') \
                .grid(row=2, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'on the basis of ability to explain the subject ? ') \
                .grid(row=3, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'with respect to his/her commitment/sincerity ? ') \
                .grid(row=4, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'on the basis of ability to generate interest'
                                             ' about the subject ? ') \
                .grid(row=5, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'with respect to his/her participation in QA sessions '
                                             'with the students ? ') \
                .grid(row=6, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'on the basis of his/her availability to the students ? ') \
                .grid(row=7, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'with respect to his/her consistency in '
                                             'conducting tests/viva and giving projects ? ') \
                .grid(row=8, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='How do you rate the faculty ' +
                                             'with respect to his/her transperency in '
                                             'evaluation of student performances ? ') \
                .grid(row=9, column=0, sticky='W', padx=12, pady=6)

            ttk.Label(self.tabs_arr[i], text='Give an Overall rating for the faculty' + ' :- ') \
                .grid(row=10, column=0, sticky='W', padx=12, pady=6)

            # ------------------------------ Adding the RadioButtons ---------------------------

            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_knowledge[i],
                                value=j + 1).grid(column=j + 1, row=1, sticky='W', padx=4)
            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_comm_skill[i],
                                value=j + 1).grid(column=j + 1, row=2, sticky='W', padx=4)
            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_explain[i],
                                value=j + 1).grid(column=j + 1, row=3, sticky='W', padx=4)
            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_commitment[i],
                                value=j + 1).grid(column=j + 1, row=4, sticky='W', padx=4)
            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_interest_gen[i],
                                value=j + 1).grid(column=j + 1, row=5, sticky='W', padx=4)

            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_qa[i],
                                value=j + 1).grid(column=j + 1, row=6, sticky='W', padx=4)

            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_accessibility[i],
                                value=j + 1).grid(column=j + 1, row=7, sticky='W', padx=4)

            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_tests[i],
                                value=j + 1).grid(column=j + 1, row=8, sticky='W', padx=4)

            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_evaluation[i],
                                value=j + 1).grid(column=j + 1, row=9, sticky='W', padx=4)

            for j in range(5):
                ttk.Radiobutton(self.tabs_arr[i], text=str(j + 1), variable=self.rating_overall[i],
                                value=j + 1).grid(column=j + 1, row=10, sticky='W', padx=4)

        txt = ['PAPER\nCODE', 'ASSIGNED\nFACULTY ', 'KNOWLEDGE',
               'COMMUNICATION\nSKILL', 'ABILITY TO\nEXPLAIN',
               'SINCERITY', 'INTEREST\nGENERATION', 'Q&A\nSESSIONS',
               'AVAILABILITY', 'CONDUCTING \nTESTS', 'TRAPERENCY IN\nEVALUATION OF\nSTUDENT PERFORMANCES',
               'OVERALL\nRATING']

        for i in range(len(txt)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=txt[i]) \
                .grid(row=0, column=i, sticky="W", padx=6, pady=4)

        ttk.Button(self.tabs_arr[len(self.tabs_arr) - 1], text="Current\nResponse", width=15,
                   command=self.current_response) \
            .grid(row=len(self.faculty) + 1, column=12, sticky="W")

    def _quit(self):
        self.window.quit()
        self.window.destroy()
        exit()

    def about(self):
        m_box.showinfo('ABOUT ', 'content.....')

    def current_response(self):

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1],
                     text=self.paper_code[i]). \
                grid(row=i + 1, column=0, sticky="w", padx=6, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=self.faculty[i]). \
                grid(row=i + 1, column=1, sticky="w", padx=6, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_knowledge[i].get())) \
                .grid(row=i + 1, column=2, sticky="w", padx=14, pady=2)
        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_comm_skill[i].get())) \
                .grid(row=i + 1, column=3, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_explain[i].get())) \
                .grid(row=i + 1, column=4, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_commitment[i].get())) \
                .grid(row=i + 1, column=5, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_interest_gen[i].get())) \
                .grid(row=i + 1, column=6, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_qa[i].get())) \
                .grid(row=i + 1, column=7, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_accessibility[i].get())) \
                .grid(row=i + 1, column=8, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_tests[i].get())) \
                .grid(row=i + 1, column=9, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_evaluation[i].get())) \
                .grid(row=i + 1, column=10, sticky="w", padx=14, pady=2)

        for i in range(len(self.paper_code)):
            tk.Label(self.tabs_arr[len(self.tabs_arr) - 1], text=str(self.rating_overall[i].get())) \
                .grid(row=i + 1, column=11, sticky="w", padx=14, pady=2)

        ttk.Button(self.tabs_arr[len(self.tabs_arr) - 1], text='SUBMIT', width=20, command=self.submit_response,
                   style="Submit.TButton") \
            .grid(row=len(self.faculty) + 2, column=6, sticky="W")

    def submit_response(self):
        choice = m_box.askyesno("Title", "Are you sure ?")
        if choice is True:
            self.results()
            m_box.showinfo("Title", "THANKS " + self.student_name + " !!!")
            m_box.showinfo('title', "BYE!!")
            self._quit()
        else:
            pass

    def results(self):
        res = []
        for i in range(len(self.paper_code)):
            temp = (self.paper_code[i], self.paper_name[i], self.faculty[i], self.student_year,
                    self.student_section, self.rating_knowledge[i].get(), self.rating_comm_skill[i].get(),
                    self.rating_explain[i].get(), self.rating_commitment[i].get(), self.rating_interest_gen[i].get(),
                    self.rating_qa[i].get(), self.rating_accessibility[i].get(), self.rating_tests[i].get(),
                    self.rating_evaluation[i].get(), self.rating_overall[i].get())
            res.append(temp)
        self.submit(self.student_roll, res)

    def submit(self, password, result):
        try:
            for i in range(len(result)):
                cursor.execute("insert into results(paper_code,paper_name,faculty,year,section,knowledge,\
                               comm,explain_ability,commitment,interest,qa,availability,tests,evaluation,\
                               overall) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                               (result[i][0], result[i][1], result[i][2], result[i][3], result[i][4],
                                result[i][5], result[i][6], result[i][7], result[i][8], result[i][9],
                                result[i][10], result[i][11], result[i][12], result[i][13], result[i][14]))
                conn.commit()

            cursor.execute("UPDATE student_data SET status=%s WHERE roll=%s", ("INELIGIBLE", password))
            conn.commit()
        except:
            pass
        conn.close()


class LoginFaculty:
    def __init__(self, this_window, previous_window):
        self.win = this_window
        self.prev_win = previous_window
        self.win.title("FACULTY LOGIN PAGE")
        self.win.geometry('+550+250')
        self.win.resizable(False, False)

        # ------------- Get faculty details ---------------
        cursor.execute('select * from faculty_login')
        self.faculty_detail = cursor.fetchall()

        # ------ Font and Style -------

        self.customized_font = font.Font(family='Helvetica', size=14, weight="bold", slant='italic')

        self.style = ttk.Style(self.win)
        self.style.configure("labels.TLabel", foreground="green", font=('Helvetica', 12))
        self.style.configure("entry.TEntry", foreground='#ff2758')
        self.style.configure("buttons.TButton", foreground="red", background='#590731',
                             font=('Helvetica', 11, 'bold'))
        self.style.configure('buttons1.TButton', foreground='black', background='#4f2e68',
                             font=('Helvetica', 10, 'bold', 'italic'))

        # Creating a Menu Bar

        self.menu_bar = Menu(self.win)

        self.win.config(menu=self.menu_bar)

        self.file_menu = Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Exit", command=self._quit)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        self.help_menu = Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="About", command=self.about)
        self.menu_bar.add_cascade(label="Help", menu=self.help_menu)

        # Creating Label frames

        self.labelframe1 = tk.LabelFrame(self.win, bd=0)
        self.labelframe1.pack(fill=BOTH, expand=1)

        self.labelframe2 = tk.LabelFrame(self.win, bd=0)
        self.labelframe2.pack(fill=BOTH, expand=1)

        # Creating Labels

        self.username = ttk.Label(self.labelframe1, text="USERNAME  : ", font=self.customized_font)
        self.username.grid(row=0, column=0, padx=25, pady=10, sticky="W")

        self.password = ttk.Label(self.labelframe1, text="PASSWORD : ", font=self.customized_font)
        self.password.grid(row=1, column=0, padx=25, pady=10, sticky="W")

        # Creating EntryFields

        self.name_entry = ttk.Entry(self.labelframe1, width=30, style='entry.TEntry')
        self.name_entry.grid(row=0, column=1, pady=10)

        self.password_entry = ttk.Entry(self.labelframe1, width=30, show='*', style='entry.TEntry')
        self.password_entry.grid(row=1, column=1, pady=15)

        # Creating Buttons

        self.login_btn = ttk.Button(self.labelframe2, text="LOGIN ", width=12, style='buttons.TButton')
        self.login_btn.grid(row=0, column=0, sticky='we', padx=160, pady=25)

        self.go_to_homepage = ttk.Button(self.labelframe2, text="Home page \n    <<---",
                                         width=14, style='buttons1.TButton')
        self.go_to_homepage.grid(row=1, column=0, sticky='w', padx=20, pady=10)

        # Event Bind to the buttons
        self.login_btn.bind('<Button-1>', self.login_for_faculty)
        self.login_btn.bind('<Return>', self.login_for_faculty)
        self.go_to_homepage.bind('<Button-1>', self.prev_page)
        self.go_to_homepage.bind('<Return>', self.prev_page)

        self.name_entry.focus()

    def _quit(self):
        self.win.quit()
        self.win.destroy()
        exit()

    def about(self):
        m_box.showinfo('ABOUT ', 'content.....')

    def login_for_faculty(self, event):
        if len(self.name_entry.get()) == 0 or len(self.password_entry.get()) == 0:
            m_box.showerror('title....', 'Fill the form correctly')
        else:
            index = -1
            for i in range(len(self.faculty_detail)):
                if self.name_entry.get().upper().strip() == self.faculty_detail[i][0] and \
                        self.password_entry.get().strip() == self.faculty_detail[i][1]:
                    index = i
                    break
            if index == -1:
                m_box.showerror('', 'Not Found')
            else:
                self.prev_win.destroy()
                self.win.destroy()
                root = tk.Tk()
                ShowResults(root, self.faculty_detail[index])

    def prev_page(self, event):
        try:
            self.win.destroy()
            self.prev_win.deiconify()
        except tk.TclError:
            pass


class ShowResults:
    def __init__(self, this_window, logged_in_faculty):
        self.window = this_window
        self.window.resizable(False, False)
        self.window.geometry('+500+250')

        self.logged_in_faculty = logged_in_faculty

        # ------ Font and Style -------

        self.customized_font = font.Font(family='Helvetica', size=11, weight="bold", slant='italic')

        self.style = ttk.Style(self.window)

        self.style.configure("labels.TLabel", foreground="green",
                             font=('Helvetica', 16, 'bold', 'italic'))
        self.style.configure("buttons.TButton", foreground="red", background='#590731',
                             font=('Helvetica', 10, 'bold'))

        try:
            if self.logged_in_faculty[2] == "YES":
                cursor.execute("select * from results")
                details = cursor.fetchall()
            else:
                cursor.execute("select * from results where faculty=%s", (self.logged_in_faculty[0],))
                details = cursor.fetchall()

            self.faculties = []
            if len(details) == 0:
                self.window.iconify()
                m_box.showerror("", 'No DATA found.\n\nPlease try later.')
                self.window.destroy()

            else:
                for i in range(len(details)):
                    self.faculties.append(details[i][2])

                self.faculties = list(set(self.faculties))

                # Creating a Menu Bar

                self.menu_bar = Menu(self.window)

                self.window.config(menu=self.menu_bar)

                self.file_menu = Menu(self.menu_bar, tearoff=0)
                self.file_menu.add_command(label="Exit", command=self._quit)
                self.menu_bar.add_cascade(label="File", menu=self.file_menu)

                # --------------------- Adding Labels  --------------------------------

                self.label_name = ttk.Label(self.window, text='FACULTY : ', style='labels.TLabel')
                self.label_name.grid(row=0, column=0, sticky="W", padx=16, pady=24)

                self.label_year = ttk.Label(self.window, text='YEAR : ', style='labels.TLabel')
                self.label_year.grid(row=1, column=0, sticky="W", padx=16, pady=8)

                # -------------- Adding ComboBoxes --------------------------

                self.faculty_selected = ttk.Combobox(self.window, width=45, state='readonly')
                self.faculty_selected['values'] = self.faculties
                self.faculty_selected.grid(row=0, column=1, sticky='W', padx=16, pady=24)
                self.faculty_selected.current(0)

                self.year_selected = ttk.Combobox(self.window, width=25, state='readonly')
                self.year_selected['values'] = ['1ST', '2ND', '3RD', '4TH']
                self.year_selected.grid(row=1, column=1, sticky='W', padx=16, pady=8)
                self.year_selected.current(0)

                # ------------ Adding Buttons ------------------

                self.get_data_btn = ttk.Button(self.window, text="GET DATA",
                                               style='buttons.TButton')
                self.get_data_btn.grid(row=2, column=1, sticky="W", padx=40, pady=20)

                self.get_data_btn.bind('<Button-1>', self.get_data)
                self.get_data_btn.bind('<Return>', self.get_data)

                self.gen_full_data = ttk.Button(self.window, text="GENERATE ALL THE RESULTS\n"
                                                                  "           (ADMIN ONLY)")
                self.gen_full_data.grid(row=3, column=1, sticky="W", padx=5, pady=20)

                self.gen_full_data.bind('<Button-1>', self.gen_full_result)
                self.gen_full_data.bind('<Return>', self.gen_full_result)

        except:
            m_box.showerror('', 'An Error has Occured !!!')
            self.window.destroy()

    def _quit(self):
        self.window.quit()
        self.window.destroy()
        exit()

    def get_data(self, event):
        cursor.execute("select * from results where year=%s and faculty=%s",
                       (self.year_selected.get(), self.faculty_selected.get()))
        data = cursor.fetchall()

        if len(data) == 0:
            m_box.showinfo('', 'Sorry !!! No data found.')
        else:
            paper_code = []
            paper_name = []

            for i in range(len(data)):
                paper_code.append(data[i][0])
                paper_name.append(data[i][1])

            paper_name = list(set(paper_name))
            paper_code = list(set(paper_code))
            root = tk.Tk()
            PreviewResult(root, self.year_selected.get(), self.faculty_selected.get(),
                          paper_code, paper_name, data)

    def gen_full_result(self, event):
        if self.logged_in_faculty[2] == "YES":
            folder_selected = filedialog.askdirectory(title="Please select a folder")
            if folder_selected == '':
                m_box.showinfo('', 'Please select a folder to store all the excel sheets..')
            else:
                res = m_box.askokcancel('!!!', 'Generating the excel files at the following location : \n\n' +
                                        folder_selected)
                if res is True:
                    create_folder = path.join(folder_selected, f'ALL RESULTS (AS OF {str(date.today())})')

                    if path.isdir(create_folder) is False:
                        mkdir(create_folder)

                    cursor.execute("select distinct year from results")
                    y = cursor.fetchall()
                    y_arr = []
                    cursor.execute("select distinct paper_code from results")
                    p_code = cursor.fetchall()
                    p_code_arr = []

                    for i in range(len(y)):
                        y_arr.append(y[i][0])
                    for i in range(len(p_code)):
                        p_code_arr.append(p_code[i][0])

                    try:
                        for i in range(len(p_code_arr)):
                            cursor.execute("select distinct faculty from results where paper_code = %s",
                                           (p_code_arr[i],))
                            temp = cursor.fetchall()
                            faculties = []
                            for j in range(len(temp)):
                                faculties.append(temp[j][0])

                            for j in range(len(faculties)):

                                # file = r'E:\PROJECTS\Student Feedback System\all_results' + '/' + faculties[j] +\
                                #        "_" + p_code_arr[i] + ".xlsx"
                                file = path.join(create_folder, faculties[j] + '_' + p_code_arr[i] + '.xlsx')

                                if path.exists(file) is False:
                                    wb = Workbook()
                                    wb.save(file)

                                wb = load_workbook(file)
                                sheet = wb['Sheet']

                                cursor.execute(
                                    "select count(paper_code) from results where faculty = %s and paper_code=%s",
                                    (faculties[j], p_code_arr[i]))

                                number_of_participant = cursor.fetchone()[0]

                                sl_no = ['i)', 'ii)', 'iii)', "iv)", 'v)', 'vi)', 'vii)', 'viii)', 'ix)', 'x)']

                                col = ['Sl no.', 'Parameter', 'No. of 5\'s', 'No. of 4\'s', 'No. of 3\'s',
                                       'No. of 2\'s',
                                       'No. of 1\'s', 'Maximum Marks', 'Marks Obtained', 'Percentage']

                                parameters = ['Knowledge based teacher (as perceived by you)',
                                              'Communication skill of the teacher',
                                              'Ability to explain the subject (as perceived by you)',
                                              'Sincerity/Commitment of the teacher',
                                              'Intereste generated by the teacher',
                                              'Welcoming questions and comments from the students and '
                                              'answering to those by the teacher',
                                              'Accessibility of the teacher in and out of the class',
                                              'Consistency of the teacher in conducting quizzes/tests/'
                                              'assignments/viva and projects',
                                              'Transperency in evaluation of students performances',
                                              'Overall rating']

                                for k in range(len(col)):
                                    sheet.cell(row=1, column=k + 1).value = col[k]
                                sheet.row_dimensions[1].height = 20

                                # ------ Adjusting cell widths --------------

                                sheet.column_dimensions['B'].width = 80
                                sheet.column_dimensions['H'].width = 18
                                sheet.column_dimensions['I'].width = 18
                                sheet.column_dimensions['J'].width = 15
                                sheet.column_dimensions['C'].width = 12
                                sheet.column_dimensions['D'].width = 12
                                sheet.column_dimensions['E'].width = 12
                                sheet.column_dimensions['F'].width = 12
                                sheet.column_dimensions['G'].width = 12
                                sheet.column_dimensions['A'].width = 8

                                values = ['5', '4', '3', '2', '1']

                                for k in range(len(values)):
                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and knowledge=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=2, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and comm=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=3, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and explain_ability=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=4, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and commitment=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=5, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and interest=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=6, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and qa=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=7, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and availability=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=8, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and tests=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=9, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and evaluation=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=10, column=k + 3).value = cursor.fetchone()[0]

                                    cursor.execute('select count(faculty) from results '
                                                   'where paper_code=%s and faculty = %s and overall=%s',
                                                   (p_code_arr[i], faculties[j], values[k]))
                                    sheet.cell(row=11, column=k + 3).value = cursor.fetchone()[0]

                                for k in range(len(sl_no)):
                                    sheet.row_dimensions[k + 2].height = 30
                                    sheet.cell(row=k + 2, column=1).value = sl_no[k]
                                    sheet.cell(row=k + 2, column=2).value = parameters[k]
                                    sheet.cell(row=k + 2, column=8).value = number_of_participant * 5

                                    temp = sheet.cell(row=k + 2, column=3).value * 5 + \
                                           sheet.cell(row=k + 2, column=4).value * 4 + \
                                           sheet.cell(row=k + 2, column=5).value * 3 + \
                                           sheet.cell(row=k + 2, column=6).value * 2 + \
                                           sheet.cell(row=k + 2, column=7).value

                                    sheet.cell(row=k + 2, column=9).value = temp

                                for k in range(len(sl_no)):
                                    temp = float(
                                        sheet.cell(row=k + 2, column=9).value / sheet.cell(row=k + 2, column=8).value)
                                    temp = round(temp * 100, 2)
                                    sheet.cell(row=k + 2, column=10).value = temp

                                wb.save(file)
                        m_box.showinfo('DONE !!!', 'Sheet Generation Successful !!')

                    except PermissionError:
                        m_box.showwarning('', 'File already opened!!')

                    except:
                        m_box.showerror('', 'Cannot generate the Excel sheets at this moment...')
                        # self.window.destroy()

        else:
            m_box.showwarning('', 'Sorry !! You dont have admin previlages')


class PreviewResult:
    def __init__(self, this_window, year, faculty, paper_code, paper_name, details):

        self.window = this_window
        self.year = year
        self.faculty = faculty
        self.paper_code = paper_code
        self.paper_name = paper_name
        self.details = details

        self.window.resizable(False, False)
        self.window.geometry('+400+100')

        # Creating a Menu Bar

        self.menu_bar = Menu(self.window)

        self.window.config(menu=self.menu_bar)

        self.file_menu = Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Exit", command=self._quit)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)

        # Creating Label frames

        self.labelframe1 = ttk.LabelFrame(self.window)
        self.labelframe1.pack(fill=BOTH, expand=1)

        self.labelframe2 = ttk.LabelFrame(self.window)
        self.labelframe2.pack(fill=BOTH, expand=1)

        self.labelframe3 = ttk.LabelFrame(self.window)
        self.labelframe3.pack(fill=BOTH, expand=1)

        self.labelframe4 = ttk.LabelFrame(self.window)
        self.labelframe4.pack(fill=BOTH, expand=1)

        self.labelframe5 = tk.LabelFrame(self.window, bd=0)
        self.labelframe5.pack(fill=BOTH, expand=1)

        # -------------- Contains inside LabelFrame1 --------------------

        self.label1 = ttk.Label(self.labelframe1, text="FACULTY NAME : " + self.faculty)
        self.label1.grid(row=0, column=0, padx=6, pady=4, sticky="W")

        self.label2 = ttk.Label(self.labelframe1, text="YEAR : " + self.year)
        self.label2.grid(row=1, column=0, padx=6, pady=4, sticky="W")

        papers_assigned = ''

        for i in paper_code:
            papers_assigned = papers_assigned + str(i) + ", "

        papers_assigned = papers_assigned[:-2]

        self.label3 = ttk.Label(self.labelframe1, text="ASSIGNED PAPER CODE(s) : " + papers_assigned)
        self.label3.grid(row=2, column=0, padx=6, pady=4, sticky="W")

        # -------------- Contains inside LabelFrame2 --------------------

        self.label4 = ttk.Label(self.labelframe2, text="PAPER CODE SELECTED :")
        self.label4.grid(row=0, column=0, sticky="W", padx=6)

        self.label5 = ttk.Label(self.labelframe2, text="Currently selected : ")
        self.label5.grid(row=1, column=0, sticky="W", padx=6)

        self.paper_selected = ttk.Combobox(self.labelframe2, width=18, height=10, state='readonly')
        self.paper_selected['values'] = paper_code
        self.paper_selected.grid(row=0, column=1, sticky='W', padx=20)
        self.paper_selected.current(0)

        self.select_btn = ttk.Button(self.labelframe2, text="SELECT ", command=self.preview)
        self.select_btn.grid(row=0, column=2, padx=5, pady=5, sticky="W")

        # -------------- Contains inside LabelFrame3 --------------------

        # -------to get the number of students participated ----------------

        cursor.execute('select count(paper_code) from results where faculty=%s and year=%s',
                       (self.faculty, self.year))

        self.number_of_participant = cursor.fetchone()[0]
        if len(paper_code) > 1:
            self.number_of_participant = int(self.number_of_participant / len(paper_code))

        self.label6 = ttk.Label(self.labelframe3, text="No. of Students Participated (i.e rated the concerned "
                                                       "faculty) : " +
                                                       str(self.number_of_participant))
        self.label6.grid(row=0, column=0, sticky='W', padx=4, pady=4)

        self.label7 = ttk.Label(self.labelframe3, text="Max marks :- No. of students participated so far * 5")
        self.label7.grid(row=1, column=0, sticky='W', padx=4, pady=4)

        # -------------- Contains inside LabelFrame4 --------------------

        col = ['Sl no.', 'Parameter', 'Maximum Marks', 'Marks Obtained', 'Percentage']
        self.sl_no = ['i)', 'ii)', 'iii)', "iv)", 'v)', 'vi)', 'vii)', 'viii)', 'ix)', 'x)']
        parameters = ['Knowledge based teacher (as perceived by you)',
                      'Communication skill of the teacher',
                      'Ability to explain the subject (as perceived by you)',
                      'Sincerity/Commitment of the teacher',
                      'Intereste generated by the teacher',
                      'Welcoming questions and comments from the students\nand '
                      'answering to those by the teacher',
                      'Accessibility of the teacher in and out of the class',
                      'Consistency of the teacher in conducting quizzes/tests/\n'
                      'assignments/viva and projects',
                      'Transperency in evaluation of students performances',
                      'Overall rating']

        for i in range(len(col)):
            ttk.Label(self.labelframe4, text=col[i]).grid(row=0, column=i, padx=4, pady=4, sticky="W")

        for i in range(10):
            ttk.Label(self.labelframe4, text=self.sl_no[i]).grid(row=i + 1, column=0, padx=4, pady=4, sticky="W")
            ttk.Label(self.labelframe4, text=parameters[i]).grid(row=i + 1, column=1, padx=4, pady=4, sticky="W")

        # -------------- Contains inside LabelFrame4 --------------------
        ttk.Button(self.labelframe5, text="Generate the full result\nin an excel sheet",
                   command=self.generate_excel_sheet).grid(row=0, column=0, pady=10, padx=250)

    def preview(self):
        self.label5.config(text="Currently selected : " + self.paper_selected.get())

        # -----Result calculation ----------

        p1_total = 0
        p2_total = 0
        p3_total = 0
        p4_total = 0
        p5_total = 0
        p6_total = 0
        p7_total = 0
        p8_total = 0
        p9_total = 0
        p10_total = 0

        for i in range(len(self.details)):
            if self.details[i][0] == self.paper_selected.get():
                p1_total = p1_total + int(self.details[i][5])
                p2_total = p2_total + int(self.details[i][6])
                p3_total = p3_total + int(self.details[i][7])
                p4_total = p4_total + int(self.details[i][8])
                p5_total = p5_total + int(self.details[i][9])
                p6_total = p6_total + int(self.details[i][10])
                p7_total = p7_total + int(self.details[i][11])
                p8_total = p8_total + int(self.details[i][12])
                p9_total = p9_total + int(self.details[i][13])
                p10_total = p10_total + int(self.details[i][14])

        p_all = [p1_total, p2_total, p3_total, p4_total, p5_total, p6_total,
                 p7_total, p8_total, p9_total, p10_total]

        p_percentages = []

        for i in range(len(p_all)):
            p_percentages.append(round(float((p_all[i] / (self.number_of_participant * 5)) * 100), 2))

        for i in range(10):
            ttk.Label(self.labelframe4, text="          " + str(self.number_of_participant * 5)) \
                .grid(row=i + 1, column=2, padx=4, pady=4, sticky="W")
            ttk.Label(self.labelframe4, text="          " + str(p_all[i])) \
                .grid(row=i + 1, column=3, padx=4, pady=4, sticky="W")
            ttk.Label(self.labelframe4, text="     " + str(p_percentages[i])) \
                .grid(row=i + 1, column=4, padx=4, pady=4, sticky="W")

    def generate_excel_sheet(self):

        folder_selected = filedialog.askdirectory(title="Please select a folder")
        if folder_selected == '':
            m_box.showinfo('', 'Please select a folder to store the excel sheets..')
        else:
            res = m_box.askokcancel('!!!', 'Generating the excel files at the following location : \n\n      ' +
                                    folder_selected)

            if res is True:
                try:
                    for i in range(len(self.paper_code)):
                        # file = folder_selected + '/' + self.faculty + "_" + self.paper_code[i] + ".xlsx"
                        file = path.join(folder_selected, self.faculty + "_" + self.paper_code[i] + ".xlsx")

                        if path.exists(file) is False:
                            wb = Workbook()
                            wb.save(file)

                        wb = load_workbook(file)
                        sheet = wb['Sheet']

                        col = ['Sl no.', 'Parameter', 'No. of 5\'s', 'No. of 4\'s', 'No. of 3\'s', 'No. of 2\'s',
                               'No. of 1\'s', 'Maximum Marks', 'Marks Obtained', 'Percentage']

                        parameters = ['Knowledge based teacher (as perceived by you)',
                                      'Communication skill of the teacher',
                                      'Ability to explain the subject (as perceived by you)',
                                      'Sincerity/Commitment of the teacher',
                                      'Intereste generated by the teacher',
                                      'Welcoming questions and comments from the students and '
                                      'answering to those by the teacher',
                                      'Accessibility of the teacher in and out of the class',
                                      'Consistency of the teacher in conducting quizzes/tests/'
                                      'assignments/viva and projects',
                                      'Transperency in evaluation of students performances',
                                      'Overall rating']

                        for j in range(len(col)):
                            sheet.cell(row=1, column=j + 1).value = col[j]
                        sheet.row_dimensions[1].height = 20

                        # ------ Adjusting cell widths --------------

                        sheet.column_dimensions['B'].width = 80
                        sheet.column_dimensions['H'].width = 18
                        sheet.column_dimensions['I'].width = 18
                        sheet.column_dimensions['J'].width = 15
                        sheet.column_dimensions['C'].width = 12
                        sheet.column_dimensions['D'].width = 12
                        sheet.column_dimensions['E'].width = 12
                        sheet.column_dimensions['F'].width = 12
                        sheet.column_dimensions['G'].width = 12
                        sheet.column_dimensions['A'].width = 8

                        values = ['5', '4', '3', '2', '1']

                        for j in range(len(values)):
                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and knowledge=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=2, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and comm=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=3, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and explain_ability=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=4, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and commitment=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=5, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and interest=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=6, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and qa=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=7, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and availability=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=8, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and tests=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=9, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and evaluation=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=10, column=j + 3).value = cursor.fetchone()[0]

                            cursor.execute('select count(faculty) from results '
                                           'where paper_code=%s and faculty = %s and overall=%s',
                                           (self.paper_code[i], self.faculty, values[j]))
                            sheet.cell(row=11, column=j + 3).value = cursor.fetchone()[0]

                        for j in range(len(self.sl_no)):
                            sheet.row_dimensions[j + 2].height = 30
                            sheet.cell(row=j + 2, column=1).value = self.sl_no[j]
                            sheet.cell(row=j + 2, column=2).value = parameters[j]
                            sheet.cell(row=j + 2, column=8).value = self.number_of_participant * 5

                            temp = sheet.cell(row=j + 2, column=3).value * 5 + \
                                   sheet.cell(row=j + 2, column=4).value * 4 + \
                                   sheet.cell(row=j + 2, column=5).value * 3 + \
                                   sheet.cell(row=j + 2, column=6).value * 2 + \
                                   sheet.cell(row=j + 2, column=7).value

                            sheet.cell(row=j + 2, column=9).value = temp

                        for j in range(len(self.sl_no)):
                            temp = float(sheet.cell(row=j + 2, column=9).value / sheet.cell(row=j + 2, column=8).value)
                            temp = round(temp * 100, 2)
                            sheet.cell(row=j + 2, column=10).value = temp

                        wb.save(file)
                    m_box.showinfo('DONE !!!', 'Sheet Generation Successful !!')
                    self.window.destroy()

                except PermissionError:
                    m_box.showwarning('', 'File already opened!!')

                except:
                    m_box.showerror('', 'Cannot generate the Excel sheets at this moment...')
                    self.window.destroy()

    def _quit(self):
        self.window.quit()
        self.window.destroy()
        exit()


def main():
    start = tk.Tk()
    Welcome(start)
    start.mainloop()


if __name__ == '__main__':
    try:
        conn = mysql.connect(**admin_credentials.dbcredentials, database="cse_database")
        cursor = conn.cursor()
        cursor.execute("select * from student_data")
        rows = cursor.fetchall()
        main()
    except mysql.errors.InterfaceError:
        win = tk.Tk()
        win.iconify()
        m_box.showerror("ERROR !!", "COULD NOT CONNECT TO THE DATABASE !!!")
